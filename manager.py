from tkinter import *
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import sqlite3
from werkzeug.security import generate_password_hash,check_password_hash
from datetime import datetime,timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
import pandas as pd
import numpy as np


class Data_base():

    """
    Classe para criação e pesquisar no banco de dados
    """

    def open_sql(self):
        """
        Abrir banco de dados e cria cursor
        """
        self.conn = sqlite3.connect("database/luxurywheels.db")
        self.cursor = self.conn.cursor()

    def close_sql(self):
        """
        Fecha banco de dados
        """
        self.conn.close()

    def table_admin(self):
        """
        Criação da tabela administrador para login do app
        """
        self.open_sql()
        admin_table =   """
                        CREATE TABLE IF NOT EXISTS administrators (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name      varchar(30) NOT NULL,
                        password  varchar(50) NOT NULL
                        );
                        """
        self.cursor.execute(admin_table)
        self.conn.commit()
        self.close_sql()

    def table_iva(self):
        """
        Criação da tabela de pagamentos dos veiculos com iva pagos
        """
        self.open_sql()

        iva_table =   """
                        CREATE TABLE IF NOT EXISTS iva_payments (
                        id	INTEGER NOT NULL,
                        name	varchar(30) NOT NULL,
                        type	varchar(50) NOT NULL,
                        price	INTEGER NOT NULL,
                        date    DATE NOT NULL,
                        PRIMARY KEY("id")
                                            );
                        """
        self.cursor.execute(iva_table)
        self.conn.commit()
        self.close_sql()

    def query_admin(self):
        """
        :return: Pesquisa na tabela administrators e retorna todos os resultados
        """
        self.open_sql()
        verify_table_admin = """SELECT * FROM  administrators"""
        result = self.cursor.execute(verify_table_admin)
        return result.fetchall()

    def create_admin(self):
        """
        Criação do login do administrador para o app
        """
        self.open_sql()
        self.query_admin()
        if self.query_admin() == []:

            pass_cryp = generate_password_hash('admin', method='pbkdf2:sha256')
            admin= [1,"admin",pass_cryp]
            admin_login = """INSERT INTO administrators VALUES(?,?,?);"""
            self.cursor.execute(admin_login,admin)
            self.conn.commit()

        self.close_sql()

    def query_type(self):
        """
        :return: Pesquisa e retorna todos o tipos de veículos
        """
        self.open_sql()
        verify_table_types = f"SELECT * FROM vehicle_types"
        result = self.cursor.execute(verify_table_types)
        return result.fetchall()

    def query_clients(self):
        """
        :return: Pesquisa e retorna todos o clientes cadastrados
        """
        self.open_sql()
        verify_clients = f"SELECT * FROM clients"
        result = self.cursor.execute(verify_clients)
        return result.fetchall()

    def query_Category(self):
        """
        :return: Pesquisa e retorna todas as categorias de clientes
        """
        self.open_sql()
        verify_table_category = f"SELECT * FROM categories"
        result = self.cursor.execute(verify_table_category)
        return result.fetchall()

    def query_vehicle(self):
        """
        Pesquisa todos os veículos cadastrados
        :return:  id, nome, status do aluguel, número de serviço, data de serviço,
        data do último iva, data do próximo iva, nome da categoria do veículo
        """
        self.open_sql()
        vehicle_table = """SELECT 
        v.id,
        v.name,
        CASE WHEN v.status =1 AND v.service = 5 THEN 'Maintenance' 
            WHEN v.status=1 THEN 'Available'
            WHEN v.status=0 THEN 'Rented'
            ELSE 'other'
       END,
        v.service,
        v.date_service,
        v.iva,
        v.next_iva,
        v.price_day,
        c.category
        from vehicles v INNER JOIN vehicle_types t on v.type_id = t.id
        INNER JOIN categories c on v.category_id = c.id
        ORDER BY c.category ASC;
        """
        result = self.cursor.execute(vehicle_table)

        return result

    def query_rents(self):
        """
        Pesquisa todos os alugueis cadastrados
        :return:  id, nome do cliente, nome do veículo, data do início do aluguel, data de término do aluguel,
        preço da diária, status do aluguel, preço total do aluguel
        """
        self.open_sql()
        rents_table = """
        SELECT 
        rents.id,
        c.name,
        v.name,
        rents.pick_up_date,
        rents.return_date,
        rents.price_day,
        CASE WHEN rents.status_rent =1 THEN 'Rented'
            WHEN rents.status_rent=0 THEN 'EndRent'
            ELSE 'other'
            END,
        rents.total_price
        FROM rents  INNER JOIN clients c on rents.client_id = c.id
        INNER JOIN vehicles v on rents.vehicle_id = v.id
        ORDER BY rents.return_date DESC;
        """
        result = self.cursor.execute(rents_table)

        return result

    def query_service(self):
        """
        Pesquisa todos os veículos com status 5 (precisam fazer revisão) que irão ficar indisponíveis por 30 dias
        :return: id, nome, statu, status do serviço, data do último iva, data do próximo iva, data da revisão
        """
        self.open_sql()

        service_table = """
        SELECT 
        v.id,
        v.name,
        CASE WHEN v.status=1 THEN 'Available'
            WHEN v.status=0 THEN 'Rented'
            ELSE 'other'
       END,
        v.service,
        v.iva,
        r.return_date,
        v.date_service
               
        
        FROM vehicles v INNER JOIN rents r on v.id = r.vehicle_id
        WHERE v.service > 4 
        GROUP BY v.name
        ORDER BY r.return_date desc ;
        """
        result = self.cursor.execute(service_table)
        return result

    def query_iva(self):
        """
        Pesquisa todos os veículos com iva em atraso para pagamento
        :return: id, tipo, nome, data do último iva, data do próximo iva
        """
        date = str(datetime.now().date())
        self.open_sql()
        vehicle_table = """SELECT 
                v.id,
                t.type,
                v.name,
                v.iva,
                v.next_iva
                
                from vehicles v INNER JOIN vehicle_types t on v.type_id = t.id
                WHERE v.next_iva <= '{}'
                ORDER BY v.iva ASC;
                """.format(date)
        result = self.cursor.execute(vehicle_table)

        return result

    def query_iva_payments(self):
        """
        Exibe todos os dados da tabela de iva_payments
        :return: todos os dados da tabela
        """
        self.open_sql()
        iva_payments_table = """SELECT * FROM iva_payments ;"""
        result = self.cursor.execute(iva_payments_table)

        return result.fetchall()

class App_admin(Data_base):

    def __init__(self):
        root = Tk()
        self.first_root = root
        self.variable()
        self.colors()
        self.screen()
        self.screen_frame_1()
        self.screen_frame_2()
        self.label_notification()
        self.table_iva()
        self.table_admin()
        self.query_admin()
        self.create_admin()
        self.query_iva_payments()
        self.functions_for_combobox()



        root.mainloop()

    def variable(self):
        """
        Armazenar todas as variáveis utilizadas na criação do app
        """
        self.var_login = StringVar()
        self.var_password = StringVar()
        self.select_type_vehicle = StringVar()
        self.select_name_vehicle = StringVar()
        self.select_service_vehicle = StringVar()
        self.select_last_iva_vehicle = StringVar()
        self.select_next_iva_vehicle = StringVar()
        self.select_price_vehicle = IntVar()
        self.select_category_vehicle = StringVar()

    def colors(self):
        """
        Cores utilizadas na criação do app
        """
        self.deepskyblue = "#00688B"
        self.deepskyblue4 = "#1C86EE"
        self.darkslategray ="#2F4F4F"
        self.brown1 = "#FF4040" #vermelho para aviso
        self.ghostwhite = "#F8F8FF"

    def screen(self):
        """
        Inicialização e configuração da janela principal do tkinter
        """
        self.first_root.title("Luxury Wheels")  # altera titulo da janela
        self.first_root.resizable(True, True)  # permite redimensionar a janela
        self.first_root.geometry("700x500")
        self.first_root.configure(background=self.deepskyblue)
        self.first_root.minsize(width=700,height=500)
        self.first_root.maxsize(width=900,height=800)
        self.first_root.iconbitmap('./static/assets/car.ico')
        self.first_root.iconify()

    def screen_frame_1(self):
        """
        Criação do frame para titulo da janela principal
        """
        self.frame_1 = Frame(self.first_root)
        self.frame_1.place(relx=0.05,rely=0.03,relwidth=0.9, relheight=0.2)

        self.label_title = ttk.Label(self.frame_1, text="Luxury Wheels", font=("times", 30, "bold"))
        self.label_title.place(relx=0.3, rely=0.2)

        global photo_title
        photo_title = PhotoImage(file="static\img\mercedesglc.png").subsample(2, 2)
        self.label_photo_title = ttk.Label(self.frame_1, image=photo_title)
        self.label_photo_title.place(relx=0.01, rely=0.03)

        self.label_photo_title_2 = ttk.Label(self.frame_1, image=photo_title)
        self.label_photo_title_2.place(relx=0.75, rely=0.03)

    def screen_frame_2(self):
        """
        Criação do frame para os campos de login e password do administador
        """
        self.frame_2 = Frame(self.first_root)
        self.frame_2.place(relx=0.05, rely=0.30, relwidth=0.9, relheight=0.6)

        self.label_login_title = ttk.LabelFrame(self.frame_2,text="Admin",relief="solid")
        self.label_login_title.place(relx=0.01,rely=0.01,width=700,height=300)

        self.login = ttk.Label(self.label_login_title,text="Login: ", font=("times",20,"bold"))
        self.login.place(relx=0.15,rely=0.2)

        self.entry_login = ttk.Entry(self.label_login_title,textvariable=self.var_login,font=("times",15))
        self.entry_login.place(relx=0.35,rely=0.21,relwidth=0.4, relheight=0.1)
        self.entry_login.focus()

        self.password = ttk.Label(self.label_login_title, text="Password: ", font=("times", 20, "bold"))
        self.password.place(relx=0.15, rely=0.4)

        self.entry_password = ttk.Entry(self.label_login_title, textvariable=self.var_password,show="*",font=(20))
        self.entry_password.place(relx=0.35, rely=0.41, relwidth=0.4, relheight=0.1)

        self.photo_login = PhotoImage(file="static/assets/butonlogin.png").subsample(10,11)
        self.button_login = ttk.Button(self.label_login_title,command=self.autentication_login, image=self.photo_login)
        self.button_login.place(relx=0.45 , rely=0.6)
        self.button_login.bind('<Return>',lambda x: x==self.autentication_login())

    def label_notification(self):
        """
        Label que exibe as informações de falha de login ou senha
        """
        self.label_login = ttk.Label(self.frame_2,text="",font=(15),background=self.brown1,foreground=self.ghostwhite)
        self.label_login.place(relx=0.5,rely=0.07)
        self.label_login.place_forget()

    def autentication_login(self):
        """
        Verificação dos campos de login e senha inseridos pelo usuário com a tabela de administrador para acessar o app
        :return: screen_frame_main_data - menu principal de pesquisa
        """
        if len(self.var_login.get()) == 0 or len(self.var_password.get()) == 0:
            self.label_login["text"] = "Login or Password does not exist"
            self.label_login.place(relx=0.4, rely=0.07)
            self.var_login.set("")
            self.var_password.set("")
        else:
            for i in self.query_admin():
                pass_cryp = check_password_hash(i[2], self.var_password.get())
                print(pass_cryp)
                if self.var_login.get() != i[1] or pass_cryp == False:
                    self.label_login["text"] = "Login ou senha inválidos"
                    self.label_login.place(relx=0.5, rely=0.07)
                    self.var_login.set("")
                    self.var_password.set("")
                else:
                    self.label_login["text"] = "Logado"
                    self.label_login.place(relx=0.5, rely=0.07)
                    self.var_login.set("")
                    self.var_password.set("")
                    return self.screen_frame_main_data()

    def screen_frame_main_data(self):
        """
        Menu principal: criação dos botões de pesquisa de veículos,alugueis,veículos em serviço,iva atrasado,
        adicionar veículo,exibir gráfico financeiro
        :return:
        """
        self.frame_1.place_forget()
        self.frame_2.place_forget()

        self.frame_3 = Frame(self.first_root,background=self.deepskyblue)
        self.frame_3.place(relx=0.01,rely=0.01,relwidth=0.98, relheight=0.99)

        self.label_main_options = ttk.Label(self.frame_3,background=self.deepskyblue)
        self.label_main_options.place(relx=0.02, rely=0.03, relwidth=0.96, relheight=0.2)

        self.label_main_view_vehicle = ttk.Label(self.frame_3,background=self.deepskyblue)
        self.label_main_view_vehicle.place(relx=0.02, rely=0.25, relwidth=0.96, relheight=0.7)

        self.button_list_vehicle = Button(self.label_main_options, text="VEHICLES",command=self.list_vehicle,
                                          font=("bold"),foreground="white",
                                          bg=self.darkslategray)
        self.button_list_vehicle.place(relx=0.01, rely=0.02, relwidth=0.2, relheight=0.25)

        self.button_add_vehicle = Button(self.label_main_options, text="NEW VEHICLE", font=("bold"),
                                           command=self.menu_add_vehicle, foreground="white",
                                           bg=self.darkslategray)
        self.button_add_vehicle.place(relx=0.01, rely=0.35, relwidth=0.2, relheight=0.25)

        self.button_list_rent = Button(self.label_main_options, text="RENTS",command=self.list_rents,
                                       font=("bold"),foreground="white",
                                          bg=self.darkslategray)
        self.button_list_rent.place(relx=0.27, rely=0.02, relwidth=0.2, relheight=0.25)

        self.button_service = Button(self.label_main_options, text="VERIFY SERVICE", font=("bold"),
                                     command=self.service_maintenance, foreground="white",
                                     bg=self.darkslategray)
        self.button_service.place(relx=0.53, rely=0.02, relwidth=0.2, relheight=0.25)


        self.button_iva = Button(self.label_main_options, text="VERIFY IVA", font=("bold"),foreground="white",
                                         command=self.list_iva, bg=self.darkslategray)
        self.button_iva.place(relx=0.79, rely=0.02, relwidth=0.2, relheight=0.25)

        self.label_expense = Label(self.label_main_options, text="Expense:", font=("bold"),
                                   foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_expense.place(relx=0.10, rely=0.75, relwidth=0.1, relheight=0.25)

        self.label_expense_value = Label(self.label_main_options,text='',anchor='w', font=("bold"),
                                         foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_expense_value.place(relx=0.21, rely=0.75, relwidth=0.15, relheight=0.25)

        self.label_profit = Label(self.label_main_options, text="Profit:", font=("bold"),
                                  foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_profit.place(relx=0.36, rely=0.75, relwidth=0.1, relheight=0.25)

        self.label_profit_value = Label(self.label_main_options,text='',anchor='w', font=("bold"),
                                        foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_profit_value.place(relx=0.45, rely=0.75, relwidth=0.15, relheight=0.25)

        self.label_total = Label(self.label_main_options, text="Total:", font=("bold"),
                                 foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_total.place(relx=0.70, rely=0.75, relwidth=0.1, relheight=0.25)

        self.label_total_value = Label(self.label_main_options,text='',anchor='w', font=("bold"),
                                       foreground=self.ghostwhite,background=self.deepskyblue)
        self.label_total_value.place(relx=0.79, rely=0.75, relwidth=0.15, relheight=0.25)

        self.button_graphics = Button(self.label_main_options, text="GRAPHIC", font=('Time',8,"bold"),foreground="white",
                                         command=self.graphics_openpy, bg=self.darkslategray)
        self.button_graphics.place(relx=0.90, rely=0.75, relwidth=0.1, relheight=0.25)

        self.profit_expense()

    def clear_checklist(self):
        """
        :return: apaga o Label da visualização das pesquisas
        """
        return self.label_main_view_vehicle.place_forget()

    def list_vehicle(self):
        """
        configuração da Treeview e widgets para exibir os dados de todos os veículos pesquisados
        Apaga os botões de start serviço e pagamento de iva
        :return:
        """

        try:
            self.button_start_service.place_forget()
        except:
            pass
        try:
            self.button_pay_iva.place_forget()
        except:
            pass

        self.label_main_view_vehicle.place(relx=0.02, rely=0.25, relwidth=0.96, relheight=0.7)
        self.check_data = ttk.Treeview(self.label_main_view_vehicle,height=3,style="mystyle.Treeview",
                                       column=("col1","col2","col3","col4","col5","col6","col7","col8",'col9'))
        self.check_data.tag_configure('bg', background='yellow')
        self.check_data.tag_configure('fg', foreground="red")
        self.check_data.tag_configure('bg2', background="red")
        self.check_data.tag_configure('fg2', foreground="black")
        self.check_data.place(relx=0.01, rely=0.02, relwidth=0.96, relheight=0.95)

        self.check_data.heading("#0", text="", anchor=CENTER)
        self.check_data.heading("#1", text="id", anchor=CENTER)
        self.check_data.heading("#2", text="Vehicle", anchor=CENTER)
        self.check_data.heading("#3", text="Status", anchor=CENTER)
        self.check_data.heading("#4", text="Service", anchor=CENTER)
        self.check_data.heading("#5", text="Date Service", anchor=CENTER)
        self.check_data.heading("#6", text="IVA", anchor=CENTER)
        self.check_data.heading("#7", text="Next IVA", anchor=CENTER)
        self.check_data.heading("#8", text="Price", anchor=CENTER)
        self.check_data.heading("#9", text="Category", anchor=CENTER)


        self.check_data.column("#0",width=1)
        self.check_data.column("#1", width=20)
        self.check_data.column("#2", width=100)
        self.check_data.column("#3", width=85)
        self.check_data.column("#4", width=60)
        self.check_data.column("#5", width=70)
        self.check_data.column("#6", width=70)
        self.check_data.column("#7", width=65)
        self.check_data.column("#8", width=70)
        self.check_data.column("#9", width=90)

        self.scrool = Scrollbar(self.label_main_view_vehicle,orient="vertical")
        self.check_data.configure(yscrollcommand=self.scrool.set)
        self.scrool.place(relx=0.96, rely=0.02, relwidth=0.05, relheight=0.95)

        self.check_data.delete(*self.check_data.get_children())

        for check_vehicle in self.query_vehicle():

            if check_vehicle[2] == 'Rented':
                self.check_data.insert("",END,values=check_vehicle,tags=('fg','bg'))
            elif check_vehicle[3] > 4:
                if check_vehicle[4] < str(datetime.now().date()):
                    self.check_data.insert("", END, values=check_vehicle, tags=('fg', 'bg'))
                else:
                    self.check_data.insert("", END, values=check_vehicle, tags=('fg', 'bg'))
            elif datetime.strptime(check_vehicle[6],'%Y-%m-%d').date() <= datetime.now().date():
                self.check_data.insert("", END, values=check_vehicle, tags=('fg','bg'))
            else:
                self.check_data.insert("", END, values=check_vehicle)

        for check_vehicle in self.query_vehicle():

            if check_vehicle[3] > 4 and check_vehicle[4] < str(datetime.now().date()):
                messagebox.showinfo("Service Car", "Vehicle need service!")
                break

        for check_vehicle in self.query_vehicle():
            if datetime.strptime(check_vehicle[6],'%Y-%m-%d').date() <= datetime.now().date():
                messagebox.showinfo("Service Car", "Vehicle need to pay IVA!")
                break
        self.check_need_buy_vehicle()
        self.close_sql()

    def menu_add_vehicle(self):
        """
        Nova janela com os widgets para adicionar um novo veícula ao banco de dados
        :return:
        """
        self.menu_add_vehicle = Toplevel()
        self.menu_add_vehicle.title("Luxury Wheels")  # altera titulo da janela
        self.menu_add_vehicle.geometry("500x500")
        self.menu_add_vehicle.configure()
        self.menu_add_vehicle.resizable(False,False)
        self.menu_add_vehicle.iconbitmap('./static/assets/car.ico')
        self.menu_add_vehicle.focus_force()
        self.menu_add_vehicle.grab_set()

        self.frame_add_vehicle = Frame(master=self.menu_add_vehicle,background=self.deepskyblue)
        self.frame_add_vehicle.place(relx=0.01, rely= 0.01,relwidth=0.98, relheight=0.98 )

        self.label_insert_text = ttk.Label(master=self.frame_add_vehicle,text="Insert new vehicle",
                                           font=("times", 25, "bold"),background=self.darkslategray,
                                           foreground=self.ghostwhite,relief='groove')
        self.label_insert_text.place(relx=0.23, rely= 0.01,relwidth=0.53, relheight=0.10 )
         # ----------------
        self.label_type =ttk.Label(master=self.frame_add_vehicle,text="Type: ",
                                           font=("times", 12, "bold"),background=self.deepskyblue,
                                           foreground=self.ghostwhite,relief='flat')
        self.label_type.place(relx=0.05, rely= 0.15,relwidth=0.10, relheight=0.05)


        self.combobox_type = ttk.Combobox(master=self.frame_add_vehicle, values=self.list_types_vehicles,
                                          font=("times", 12, "bold"), background='red',
                                              textvariable=self.select_type_vehicle,foreground=self.deepskyblue)


        self.combobox_type.place(relx=0.18, rely=0.15, relwidth=0.26, relheight=0.05)

        self.label_name = ttk.Label(master=self.frame_add_vehicle, text="Name: ",
                                    font=("times", 12, "bold"), background=self.deepskyblue,
                                    foreground=self.ghostwhite, relief='flat')
        self.label_name.place(relx=0.05, rely=0.25, relwidth=0.10, relheight=0.05)

        self.entry_name = ttk.Entry(master=self.frame_add_vehicle,textvariable=self.select_name_vehicle,
                                          font=("times", 12, "bold"),
                                          foreground='black')
        self.entry_name.place(relx=0.18, rely=0.25, relwidth=0.26, relheight=0.05)

        self.label_date_service = ttk.Label(master=self.frame_add_vehicle, text="Date Service: ",
                                            font=("times", 12, "bold"), background=self.deepskyblue,
                                            foreground=self.ghostwhite, relief='flat')
        self.label_date_service.place(relx=0.05, rely=0.35, relwidth=0.20, relheight=0.05)

        self.entry_date_service = DateEntry(master=self.frame_add_vehicle, showweeknumbers=False,
                                            textvariable=self.select_service_vehicle,
                                            showothermonthdays=False, weekendbackground=self.deepskyblue,
                                            weekendforeground='white', font=("times", 10, "bold"),
                                            background=self.darkslategray, normalbackground=self.deepskyblue,
                                            foreground=self.ghostwhite,date_pattern='y-mm-dd')
        self.entry_date_service.place(relx=0.26, rely=0.35, relwidth=0.18, relheight=0.05)

        self.label_last_iva = ttk.Label(master=self.frame_add_vehicle, text="Last Iva: ",
                                    font=("times", 12, "bold"), background=self.deepskyblue,
                                    foreground=self.ghostwhite, relief='flat')
        self.label_last_iva.place(relx=0.60, rely=0.15, relwidth=0.14, relheight=0.05)

        self.entry_last_iva = DateEntry(master=self.frame_add_vehicle,showweeknumbers=False,
                                            showothermonthdays = False,weekendbackground=self.deepskyblue,
                                        textvariable=self.select_last_iva_vehicle,
                                            weekendforeground='white',font=("times", 10, "bold"),
                                            background=self.darkslategray,normalbackground=self.deepskyblue,
                                          foreground=self.ghostwhite,date_pattern='y-mm-dd')
        self.entry_last_iva.place(relx=0.75, rely=0.15, relwidth=0.19, relheight=0.05)

        self.label_next_iva = ttk.Label(master=self.frame_add_vehicle, text="Next Iva: ",
                                    font=("times", 9, "bold"), background=self.deepskyblue,
                                    foreground=self.ghostwhite, relief='flat')
        self.label_next_iva.place(relx=0.60, rely=0.25, relwidth=0.14, relheight=0.05)

        self.entry_next_iva = DateEntry(master=self.frame_add_vehicle,showweeknumbers=False,
                                            showothermonthdays = False,weekendbackground=self.deepskyblue,
                                            textvariable=self.select_next_iva_vehicle,
                                            weekendforeground='white',font=("times", 10, "bold"),
                                            background=self.darkslategray,normalbackground=self.deepskyblue,
                                          foreground=self.ghostwhite,date_pattern='y-mm-dd')
        self.entry_next_iva.place(relx=0.75, rely=0.25, relwidth=0.19, relheight=0.05)

        self.label_price = ttk.Label(master=self.frame_add_vehicle, text="Price: ",
                                    font=("times", 12, "bold"), background=self.deepskyblue,
                                    foreground=self.ghostwhite, relief='flat')
        self.label_price.place(relx=0.60, rely=0.35, relwidth=0.14, relheight=0.05)

        self.entry_price = ttk.Entry(master=self.frame_add_vehicle,textvariable=self.select_price_vehicle,
                                    font=("times", 12, "bold"), background='red',
                                    foreground='black')
        self.entry_price.place(relx=0.77, rely=0.35, relwidth=0.17, relheight=0.05)

        self.label_category = ttk.Label(master=self.frame_add_vehicle, text="Category: ",
                                    font=("times", 12, "bold"), background=self.deepskyblue,
                                    foreground=self.ghostwhite, relief='flat')
        self.label_category.place(relx=0.05, rely=0.45, relwidth=0.15, relheight=0.05)

        self.combobox_category = ttk.Combobox(master=self.frame_add_vehicle, values=self.list_category_vehicles,
                                          font=("times", 12, "bold"), background='black',
                                              textvariable=self.select_category_vehicle,foreground=self.deepskyblue)
        self.combobox_category.place(relx=0.23, rely=0.45, relwidth=0.21, relheight=0.05)

        self.button_register_vehicle = Button(master=self.frame_add_vehicle,text='Register Vechicle', font=("bold")
                                           ,  foreground="white",command=self.register_vehicle,
                                           bg=self.darkslategray)
        self.button_register_vehicle.place(relx=0.35, rely=0.55, relwidth=0.30, relheight=0.1)

        self.label_menssage = ttk.Label(master=self.frame_add_vehicle, text="",
                                        font=("times", 15, "bold"),background=self.deepskyblue,
                                        foreground=self.brown1, relief='flat')
        self.label_menssage.place(relx=0.25, rely=0.70, relwidth=0.50, relheight=0.05)

        self.button_back_menu = Button(master=self.frame_add_vehicle, text='Back', font=("bold")
                                              , foreground="white", command=self.back_main_menu,
                                              bg=self.darkslategray)
        self.button_back_menu.place(relx=0.10, rely=0.80, relwidth=0.20, relheight=0.08)

    def functions_for_combobox(self):
        """
        método para listar os tipos de veículos e cetagoria na janela de adicionar veículos
        """

        self.list_types_vehicles = []
        self.list_category_vehicles = []
        for type in self.query_type():
            self.list_types_vehicles.append(type[1])
        for cat in self.query_Category():
            self.list_category_vehicles.append(cat[1])

    def register_vehicle(self):
        """
        Verificação dos dados inseridos pelo usuário para criação do veículo
        """

        try:
            if (self.select_type_vehicle.get() == '' or self.select_name_vehicle.get() == '' or
                    self.select_category_vehicle.get() == '' or self.select_price_vehicle.get()==''):
                self.label_menssage['text'] = 'Please, enter all the fields!'
                self.select_type_vehicle.set('')
                self.select_name_vehicle.set('')
                self.select_category_vehicle.set('')
                self.select_price_vehicle.set(0)
            else:

                if self.select_type_vehicle.get() == 'Car':
                    variable_type_vehicle = 1
                else:
                    variable_type_vehicle = 2

                if self.select_category_vehicle.get() == 'Gold':
                    variable_category_vehicle = 1
                elif self.select_category_vehicle.get() == 'Silver':
                    variable_category_vehicle = 2
                else:
                    variable_category_vehicle = 3

                self.label_menssage['text'] = ''
                var_image_table_vehicles = (self.select_name_vehicle.get().strip().replace(' ','')) + '.jpg'
                print(var_image_table_vehicles)

                self.open_sql()
                new_vehicle = (f"INSERT INTO vehicles (type_id,name,status,service,date_service,iva,"
                               f"next_iva,price_day,category_id,image_car) VALUES (?,?,?,?,?,?,?,?,?,?)")
                self.cursor.execute(new_vehicle,(variable_type_vehicle,self.select_name_vehicle.get(),
                                                 True,0,self.select_service_vehicle.get(),
                                                 self.select_last_iva_vehicle.get(),self.select_next_iva_vehicle.get(),
                                                 self.select_price_vehicle.get(),variable_category_vehicle,
                                                 var_image_table_vehicles))
                self.conn.commit()
                self.close_sql()
                self.back_main_menu()
                messagebox.showinfo("Register Car", "New vehicle successfully!")

        except:
            self.label_menssage['text'] = 'Please, enter all the fields!'

    def back_main_menu(self):
        """
        método para adicionar comando ao Button "Back" para fechar a janela de adicionar veículos
        """
        self.menu_add_vehicle.destroy()

    def check_need_buy_vehicle(self):
        """
        Método para verificar a quantidade de clientes cadastrados e veículos
        :return: Se não tiver disponível pelo menos 5 veículos para cada tipo de cliente, uma mensagem é exibida
        """

        vehicle_count_gold,vehicle_count_silver,vehicle_count_economy = 0,0,0
        client_count_gold, client_count_silver, client_count_economy = 0, 0, 0
        vehicles_count_available = []

        for vehicle_available in self.query_vehicle():
            if vehicle_available[2] == 'Available':
                vehicles_count_available.append(vehicle_available)

        for vehicle in vehicles_count_available:
            vehicle_count_gold += 1
            if vehicle[8] == 'Silver' or vehicle[8] == 'Economy':
                vehicle_count_silver += 1
            if vehicle[8] == 'Economy':
                vehicle_count_economy += 1

        for client in self.query_clients():
            vehicle_count_gold += 1
            if client[4] == 1:
                client_count_gold += 1
            elif client[4] == 2:
                client_count_silver += 1
            elif client[4] == 3:
                client_count_economy += 1

        if (client_count_gold + 5) >= vehicle_count_gold:
            messagebox.showwarning('Vechile','Need to buy more vehicles!')
        elif (client_count_silver + 5) >= vehicle_count_silver:
            messagebox.showwarning('Vechile', 'Need to buy more vehicles!')
        elif (client_count_silver + 5) >= vehicle_count_economy:
            messagebox.showwarning('Vechile', 'Need to buy more vehicles!')

    def list_rents(self):
        """
        configuração da Treeview e widgets para exibir os dados de todos os alugueis pesquisados
        Apaga os botões de start serviço e pagamento de iva
        :return:
        """

        try:
            self.button_start_service.place_forget()
        except:
            pass

        try:
            self.button_pay_iva.place_forget()
        except:
            pass

        self.label_main_view_vehicle.place(relx=0.02, rely=0.25, relwidth=0.96, relheight=0.7)
        self.check_data = ttk.Treeview(self.label_main_view_vehicle,height=3,style="mystyle.Treeview",
                                       columns=("col1","col2","col3","col4","col5","col6","col7","col8"))
        self.check_data.tag_configure('bg', background='yellow')
        self.check_data.tag_configure('fg', foreground="red")
        self.check_data.place(relx=0.01, rely=0.02, relwidth=0.96, relheight=0.95)

        self.check_data.heading("#0", text="", anchor=CENTER)
        self.check_data.heading("#1", text="id", anchor=CENTER)
        self.check_data.heading("#2", text="Client", anchor=CENTER)
        self.check_data.heading("#3", text="Vehicle", anchor=CENTER)
        self.check_data.heading("#4", text="Pick_up_date", anchor=CENTER)
        self.check_data.heading("#5", text="Return_date", anchor=CENTER)
        self.check_data.heading("#6", text="Price_day", anchor=CENTER)
        self.check_data.heading("#7", text="Status_rent", anchor=CENTER)
        self.check_data.heading("#8", text="Total_price", anchor=CENTER)

        self.check_data.column("#0",width=1)
        self.check_data.column("#1", width=20)
        self.check_data.column("#2", width=90)
        self.check_data.column("#3", width=115)
        self.check_data.column("#4", width=60)
        self.check_data.column("#5", width=60)
        self.check_data.column("#6", width=60)
        self.check_data.column("#7", width=50)
        self.check_data.column("#8", width=90)

        self.scrool = Scrollbar(self.label_main_view_vehicle,orient="vertical")
        self.check_data.configure(yscrollcommand=self.scrool.set)
        self.scrool.place(relx=0.96, rely=0.02, relwidth=0.05, relheight=0.95)

        self.check_data.delete(*self.check_data.get_children())
        for check_rent in self.query_rents():

            if check_rent[6] == 'Rented':
                self.check_data.insert("",END,values=check_rent,tags=('fg','bg'))

            else:
                self.check_data.insert("", END, values=check_rent)
        self.close_sql()
        self.profit_expense()

    def service_maintenance(self):
        """
        Exibi todos os veículos que precisam fazer a revisão após 5 alugueis
        """

        try:
            self.button_pay_iva.place_forget()
        except:
            pass

        self.label_main_view_vehicle.place(relx=0.02, rely=0.25, relwidth=0.96, relheight=0.7)
        self.check_data = ttk.Treeview(self.label_main_view_vehicle, height=3, style="mystyle.Treeview",
                                       column=("col1", "col2", "col3", "col4", "col5", "col6", "col7"))
        self.check_data.tag_configure('bg', background='yellow')
        self.check_data.tag_configure('fg', foreground="red")
        self.check_data.place(relx=0.01, rely=0.02, relwidth=0.96, relheight=0.95)

        self.button_start_service = Button(self.label_main_options, text="START SERVICE", font=("bold")
                                    ,command=self.start_maintenance, foreground="white",
                                     bg=self.darkslategray)
        self.button_start_service.place(relx=0.53, rely=0.35, relwidth=0.2, relheight=0.25)

        self.check_data.heading("#0", text="", anchor=CENTER)
        self.check_data.heading("#1", text="id", anchor=CENTER)
        self.check_data.heading("#2", text="Name", anchor=CENTER)
        self.check_data.heading("#3", text="Status", anchor=CENTER)
        self.check_data.heading("#4", text="Service", anchor=CENTER)
        self.check_data.heading("#5", text="IVA", anchor=CENTER)
        self.check_data.heading("#6", text="Return_date", anchor=CENTER)
        self.check_data.heading("#7", text="Date Service", anchor=CENTER)

        self.check_data.column("#0", width=1)
        self.check_data.column("#1", width=20)
        self.check_data.column("#2", width=90)
        self.check_data.column("#3", width=115)
        self.check_data.column("#4", width=60)
        self.check_data.column("#5", width=60)
        self.check_data.column("#6", width=60)
        self.check_data.column("#7", width=70)

        self.scrool = Scrollbar(self.label_main_view_vehicle, orient="vertical")
        self.check_data.configure(yscrollcommand=self.scrool.set)
        self.scrool.place(relx=0.96, rely=0.02, relwidth=0.05, relheight=0.95)

        self.check_data.delete(*self.check_data.get_children())
        date_now = datetime.now().date()
        for check_rent in self.query_service():
            if check_rent[6] < str(date_now):
                self.check_data.insert("", END, values=check_rent, tags=('fg', 'bg'))

        if self.check_data.get_children() == ():
            self.button_start_service.place_forget()
            messagebox.showinfo("Service Car", "All vehicles service conclued!")

        self.close_sql()

    def start_maintenance(self):
        """
        Widget que verifica os veículos listados em service_maintenance e atualiza o banco de dados com a data
        de revisão (date_service) para 30 dias a frente
        """
        self.check_data.delete(*self.check_data.get_children())
        date_maintenance = timedelta(days=30)
        list_maintenance = []
        for check_rent in self.query_service():
            list_maintenance.append(check_rent)
        
        for maintenance in list_maintenance:
            if maintenance[6] < str(datetime.now().date()):
                convert_for_date = datetime.strptime(maintenance[5],'%Y-%m-%d').date()
                new_date_finish_maintenance = str(convert_for_date + date_maintenance)
                date_finish_maintenance = (f"UPDATE vehicles SET date_service = '{new_date_finish_maintenance}'"
                                           f"WHERE name = '{maintenance[1]}';")
                self.cursor.execute(date_finish_maintenance)
                self.conn.commit()

        self.button_start_service.place_forget()
        self.list_vehicle()
        self.close_sql()

    def list_iva(self):
        """
        Exibi todos os veículos que precisam fazer o pagamento de iva
        """

        try:
            self.button_start_service.place_forget()
        except:
            pass

        self.label_main_view_vehicle.place(relx=0.02, rely=0.25, relwidth=0.96, relheight=0.7)
        self.check_data = ttk.Treeview(self.label_main_view_vehicle, height=3, style="mystyle.Treeview",
                                       column=("col1", "col2", "col3", "col4", "col5"))
        self.check_data.tag_configure('bg', background='yellow')
        self.check_data.tag_configure('fg', foreground="red")
        self.check_data.tag_configure('bg2', background="red")
        self.check_data.tag_configure('fg2', foreground="black")
        self.check_data.place(relx=0.01, rely=0.02, relwidth=0.96, relheight=0.95)

        self.button_pay_iva = Button(self.label_main_options, text="Pay IVA", font=("bold")
                                           , foreground="white",command=self.pay_iva,
                                           bg=self.darkslategray)
        self.button_pay_iva.place(relx=0.79, rely=0.35, relwidth=0.2, relheight=0.25)

        self.check_data.heading("#0", text="", anchor=CENTER)
        self.check_data.heading("#1", text="id", anchor=CENTER)
        self.check_data.heading("#2", text="Type", anchor=CENTER)
        self.check_data.heading("#3", text="Name", anchor=CENTER)
        self.check_data.heading("#4", text="Iva", anchor=CENTER)
        self.check_data.heading("#5", text="Next Iva", anchor=CENTER)

        self.check_data.column("#0", width=1)
        self.check_data.column("#1", width=50)
        self.check_data.column("#2", width=100)
        self.check_data.column("#3", width=100)
        self.check_data.column("#4", width=100)
        self.check_data.column("#5", width=100)

        self.scrool = Scrollbar(self.label_main_view_vehicle, orient="vertical")
        self.check_data.configure(yscrollcommand=self.scrool.set)
        self.scrool.place(relx=0.96, rely=0.02, relwidth=0.05, relheight=0.95)

        self.check_data.delete(*self.check_data.get_children())

        date_now = datetime.now().date()
        for check_iva in self.query_iva():
            date_table = datetime.strptime(check_iva[4],'%Y-%m-%d').date()

            self.check_data.insert("", END, values=check_iva, tags=('fg', 'bg'))

        if self.check_data.get_children() == ():
            self.button_pay_iva.place_forget()
            messagebox.showinfo("Service Car", "All vehicles paid!")

    def pay_iva(self):
        """
        Método que atualiza as data de iva e next_iva, além de adicionar os dados a tabela de iva_payments
        """
        self.check_data.delete(*self.check_data.get_children())
        date_yearly_iva = timedelta(days=365)
        list_iva = []
        for check_iva in self.query_iva():
            list_iva.append(check_iva)

        for vehicle in list_iva:
            date_table = datetime.strptime(vehicle[4], '%Y-%m-%d').date()
            new_date_iva = str(date_table + date_yearly_iva)

            if vehicle[1] == 'Car':
                date_pay_iva = (f"UPDATE vehicles SET iva = '{vehicle[4]}', next_iva = '{new_date_iva}'"
                                f" WHERE name = '{vehicle[2]}';")
                self.cursor.execute(date_pay_iva)
                update_iva_prices_car = """INSERT INTO iva_payments (name,type,price,date) VALUES (?,?,?,?); """
                self.cursor.execute(update_iva_prices_car,(vehicle[2],vehicle[1],250,vehicle[4]))
                self.conn.commit()
            elif vehicle[1] == 'Motorcycle':
                date_pay_iva = (f"UPDATE vehicles SET iva = '{vehicle[4]}', next_iva = '{new_date_iva}'"
                                f" WHERE name = '{vehicle[2]}';")
                self.cursor.execute(date_pay_iva)
                update_iva_prices_motorcycle = """INSERT INTO iva_payments (name,type,price,date) VALUES (?,?,?,?);"""
                self.cursor.execute(update_iva_prices_motorcycle, (vehicle[2], vehicle[1],150,vehicle[4]))
                self.conn.commit()

        self.button_pay_iva.place_forget()
        self.list_iva()
        self.close_sql()
        self.profit_expense()

    def profit_expense(self):
        """
        Criaçao dos Labels de lucro,débito e total dos rendimentos no frame do menu inicial
        """
        profit,expense = 0,0

        for check_price_profit in self.query_rents():
            if check_price_profit[6] == 'EndRent':
                profit += check_price_profit[7]

        for check_price in self.query_iva_payments():
            expense += check_price[3]
        self.label_expense_value['text'] = str(expense) + ' £'
        self.label_profit_value['text'] = str(profit) + ' £'
        total = profit - expense
        self.label_total_value['text'] = str(total) + ' £'

    def graphics_openpy(self):
        """
        Criação do gráfico financeiro da empresa
        Armazenar os dados em excel
        Trabalhar os dados com dataframe
        """

        self.graphics = Toplevel()
        self.graphics.title("Luxury Wheels")  # altera titulo da janela
        self.graphics.geometry("700x700")
        self.graphics.configure()
        self.graphics.resizable(False, False)
        self.graphics.iconbitmap('./static/assets/car.ico')
        self.graphics.focus_force()
        self.graphics.grab_set()

        self.frame_graphics = Frame(master=self.graphics, background=self.deepskyblue)
        self.frame_graphics.place(relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98)

        #Trabalhar dados com Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'graphic'
        graphic = wb['graphic']

        list_profit = []
        for rents in self.query_rents():
            month = datetime.strptime(rents[4], '%Y-%m-%d').date()
            month_format = datetime.strftime(month, '%d/%b/%Y')
            list_profit.append([month_format,rents[7]])
        for add_rents_xlsx in list_profit:
            graphic.append(add_rents_xlsx)


        list_expense = []
        for iva in self.query_iva_payments():
            month = datetime.strptime(iva[4], '%Y-%m-%d').date()
            month_format = datetime.strftime(month, '%d/%b/%Y')
            list_expense.append([month_format, (iva[3]*(-1))])

        for add_iva_xlsx in list_expense:
            graphic.append(add_iva_xlsx)

        ws.insert_rows(1)

        ws["A1"] = "Date"
        ws["B1"] = "Profit-expense"
        ws["C1"] = "Total"

        ws.insert_rows(2)

        ws["A2"] = ""
        ws["B2"] = "0"
        ws["C2"] = "0"
        number_cell = 0

        for cell in graphic.rows:
            number_cell += 1

        total = 4
        ws["C3"] = "=B3"
        while total <= number_cell:
            ws[f"C{total}"] = f"=C{total - 1}+B{total}"
            total+=1


        wb.save('profit-expense.xlsx')

        #Trabalhar dados do excel com dataframe
        df_total = pd.read_excel("profit-expense.xlsx")

        df_excel_data = df_total.groupby('Date')['Profit-expense'].sum().reset_index()

        df_excel_data['acumulado'] = df_excel_data['Profit-expense'].cumsum()

        df_excel_data.pop('Profit-expense')

        df_excel_data['Date'] = pd.to_datetime(df_excel_data['Date'])

        df_sort = df_excel_data.sort_values(by='Date')

        profit = np.array(df_sort['Date'])
        expense = np.array(df_sort['acumulado'])

        fig = plt.Figure((6,6), 65, facecolor=self.deepskyblue )
        ax = fig.add_subplot(111)

        canvas = FigureCanvasTkAgg(fig, master=self.frame_graphics)

        canvas.get_tk_widget().place(relx=0.05, rely=0.05, relwidth=0.80, relheight=0.80)

        ax.plot(profit,expense)

        ax.set_title('EARN')
        ax.set_xlabel('date')
        ax.set_ylabel('Money')

        self.screen_frame_main_data()





